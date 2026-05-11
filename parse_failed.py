"""
Парсинг конкретных неуспешных проектов Planeta.ru.

Скрипт берёт список URL из Excel-файлов (гиперссылки в колонке
«Ссылка на страницу проекта» где «Успех» == 0), парсит каждый
через Selenium и сохраняет в wayback_class0.csv.

Запуск:
    python parse_failed.py

Можно передать несколько Excel-файлов — скрипт объединит URL
и удалит дубликаты автоматически.
"""

import time
import random
import logging
import re
import csv
from pathlib import Path
from dataclasses import dataclass, asdict
from typing import Optional

import pandas as pd
from openpyxl import load_workbook
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from webdriver_manager.chrome import ChromeDriverManager

# ─── Логгирование ─────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)s  %(message)s",
    handlers=[
        logging.FileHandler("parse_failed.log", encoding="utf-8"),
        logging.StreamHandler(),
    ],
)
log = logging.getLogger(__name__)

# ─── Конфигурация ─────────────────────────────────────────────────────────
# Список Excel-файлов с неуспешными проектами.
# Добавляйте новые файлы сюда — скрипт объединит все URL автоматически.
EXCEL_FILES = [
    "dataset.xlsx",
    "DAtaset_po_kraudfandingu.xlsx",
    "Датасет _Крауд-проекты_Исходник.xlsx"
]

# Настройки колонок (если в других файлах названия отличаются — поменяйте)
URL_COLUMN     = "Ссылка на страницу проекта"
SUCCESS_COLUMN = "Успех"
SUCCESS_VALUE  = 0   # значение «неуспешный»

OUTPUT_FILE    = "planeta_class0.csv"
DELAY_MIN      = 2.5
DELAY_MAX      = 4.5
PAGE_TIMEOUT   = 60
JS_RENDER_WAIT = 5.0


# ─── Структура данных ──────────────────────────────────────────────────────
@dataclass
class Project:
    project_id:      Optional[str]   = None
    url:             Optional[str]   = None
    title:           Optional[str]   = None
    category:        Optional[str]   = None
    status:          Optional[str]   = None
    region:          Optional[str]   = None
    goal_sum:        Optional[float] = None
    collected_sum:   Optional[float] = None
    collected_pct:   Optional[float] = None
    backers_count:   Optional[int]   = None
    # duration_days убран: вычисляется в feature_engineering.py как date_end - date_start
    nmb_video:       Optional[int]   = None
    nmb_image:       Optional[int]   = None
    nmb_reward:      Optional[int]   = None
    min_reward:      Optional[float] = None
    max_reward:      Optional[float] = None
    nmb_word:        Optional[int]   = None
    nmb_sentence:    Optional[int]   = None
    short_desc:      Optional[str]   = None
    full_text:       Optional[str]   = None   # полный текст описания (для TF-IDF, BERT)
    author_name:     Optional[str]   = None
    author_projects: Optional[int]   = None
    author_backers:  Optional[int]   = None
    date_start:      Optional[str]   = None
    date_end:        Optional[str]   = None
    updates_count:   Optional[int]   = None
    comments_count:  Optional[int]   = None
    parsed_date:     Optional[str]   = None
    target:          int             = 0    # всегда 0 для неуспешных


# ─── Извлечение URL из Excel ───────────────────────────────────────────────
def extract_failed_urls(excel_files: list[str]) -> list[str]:
    """
    Читает гиперссылки из колонки URL_COLUMN для строк где SUCCESS_COLUMN == SUCCESS_VALUE.
    Объединяет все файлы и возвращает дедуплицированный список URL.
    """
    all_urls: dict[str, str] = {}  # url -> название (для лога)

    for fname in excel_files:
        path = Path(fname)
        if not path.exists():
            log.warning("Файл не найден: %s", path.absolute())
            continue

        log.info("Читаем: %s", fname)
        try:
            wb = load_workbook(path)
        except Exception as e:
            log.error("Не удалось открыть %s: %s", fname, e)
            continue

        ws = wb.active
        headers = [cell.value for cell in ws[1]]

        if URL_COLUMN not in headers:
            log.error("Колонка '%s' не найдена в %s. Есть: %s",
                      URL_COLUMN, fname, headers)
            continue
        if SUCCESS_COLUMN not in headers:
            log.error("Колонка '%s' не найдена в %s.", SUCCESS_COLUMN, fname)
            continue

        url_idx     = headers.index(URL_COLUMN)
        success_idx = headers.index(SUCCESS_COLUMN)
        name_idx    = headers.index("Название проекта") if "Название проекта" in headers else url_idx

        found_in_file = 0
        for row in ws.iter_rows(min_row=2):
            try:
                success_val = row[success_idx].value
                if success_val is None:
                    continue
                if int(success_val) != SUCCESS_VALUE:
                    continue
            except (ValueError, TypeError):
                continue

            cell = row[url_idx]
            href = cell.hyperlink.target if cell.hyperlink else None
            name = row[name_idx].value or ""

            if href and href not in all_urls:
                all_urls[href] = name
                found_in_file += 1
            elif not href:
                log.warning("  Нет гиперссылки для: %s", name)

        log.info("Найдено новых URL: %d", found_in_file)

    urls = list(all_urls.keys())
    log.info("Итого уникальных URL неуспешных проектов: %d", len(urls))
    for i, url in enumerate(urls, 1):
        log.debug("  %d. %s", i, url)
    return urls


# ─── Браузер ───────────────────────────────────────────────────────────────
def make_driver() -> webdriver.Chrome:
    opts = Options()
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option("useAutomationExtension", False)
    opts.add_argument(
        "user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    )
    opts.add_argument("--lang=ru-RU,ru;q=0.9")
    opts.add_argument("--window-size=1400,900")
    service = Service(ChromeDriverManager().install())
    driver  = webdriver.Chrome(service=service, options=opts)
    driver.execute_cdp_cmd(
        "Page.addScriptToEvaluateOnNewDocument",
        {"source": "Object.defineProperty(navigator, 'webdriver', {get: () => undefined});"},
    )
    return driver


def load_page(driver, url: str) -> Optional[BeautifulSoup]:
    try:
        driver.get(url)
    except Exception as e:
        log.warning("Навигация: %s", e)
        return None

    deadline = time.time() + PAGE_TIMEOUT
    warned   = False

    while time.time() < deadline:
        src = driver.page_source.lower()
        if any(s in src for s in ("небольшая проверка", "just a moment", "cf-browser")):
            if not warned:
                print("\n" + "="*50)
                print("CLOUDFLARE — пройдите проверку в браузере!")
                print("="*50 + "\n")
                warned = True
            time.sleep(2)
            continue
        try:
            WebDriverWait(driver, 3).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "h1"))
            )
            time.sleep(JS_RENDER_WAIT)
            driver.execute_script("window.scrollTo(0, 600);")
            time.sleep(1.0)
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight * 0.4);")
            time.sleep(1.0)
            driver.execute_script("window.scrollTo(0, 0);")
            time.sleep(0.5)
            return BeautifulSoup(driver.page_source, "lxml")
        except TimeoutException:
            time.sleep(2)

    log.warning("Таймаут: %s", url)
    return BeautifulSoup(driver.page_source, "lxml") if driver.page_source else None


# ─── Парсинг страницы ──────────────────────────────────────────────────────
def clean_text(text: str) -> str:
    """Убирает HTML-артефакты: NBSP, нулевые пробелы, HTML-сущности."""
    if not text:
        return text
    text = text.replace('\xa0', ' ')
    text = text.replace('\u200b', '')
    text = text.replace('\u2019', "'")
    text = re.sub(r'&nbsp;', ' ', text)
    text = re.sub(r'&amp;',  '&', text)
    text = re.sub(r'&lt;',   '<', text)
    text = re.sub(r'&gt;',   '>', text)
    text = re.sub(r' {2,}', ' ', text)
    return text.strip()


def _money(text: str) -> Optional[float]:
    if not text:
        return None
    clean = re.sub(r"[^\d,.]", "", text.replace("\xa0", "").replace(" ", ""))
    try:
        return float(clean.replace(",", "."))
    except ValueError:
        return None


def _int(text: str) -> Optional[int]:
    digits = re.sub(r"[^\d]", "", text or "")
    return int(digits) if digits else None


def _words(text: str) -> int:
    return len(re.findall(r"\b\w+\b", text, re.UNICODE)) if text else 0


def _sentences(text: str) -> int:
    return len(re.split(r"[.!?]+", text.strip())) if text and text.strip() else 0




def parse_project(driver, url: str) -> Optional[Project]:
    soup = load_page(driver, url)
    if not soup:
        return None

    p = Project(url=url, target=0)
    m = re.search(r"/campaigns/([^/?#]+)", url)
    p.project_id = m.group(1) if m else url.split("/")[-1]

    # Заголовок
    h1 = soup.find("h1", class_=re.compile("title"))
    p.title = h1.get_text(strip=True) if h1 else None
    if not p.title:
        # Проверяем — может страница 404
        page_text = soup.get_text(" ", strip=True)
        if any(w in page_text.lower() for w in ("не найден", "404", "не существует")):
            log.warning("  Страница не найдена (404): %s", url)
        return None

    # Описание
    short_p = soup.find("p", class_=re.compile("description"))
    p.short_desc = short_p.get_text(strip=True) if short_p else None

    # Финансы
    sum_el = soup.find("span", class_=re.compile("fundingSumValue"))
    p.collected_sum = _money(sum_el.get_text()) if sum_el else None

    target_el = soup.find("span", class_=re.compile("fundingSumTarget"))
    p.goal_sum = _money(target_el.get_text()) if target_el else None

    pct_el = soup.find("p", class_=re.compile("progress-text"))
    if pct_el:
        p.collected_pct = _money(pct_el.get_text())
    elif p.goal_sum and p.collected_sum and p.goal_sum > 0:
        p.collected_pct = round(p.collected_sum / p.goal_sum * 100, 2)

    # Статус
    status_el = soup.find(class_=re.compile("status"))
    if status_el:
        raw = status_el.get_text(strip=True).lower()
        if any(w in raw for w in ("успешн", "завершён", "funded")):
            p.status = "successful"
        elif any(w in raw for w in ("не собрал", "failed")):
            p.status = "failed"
        elif any(w in raw for w in ("идёт", "сбор", "active")):
            p.status = "active"
    if not p.status and p.collected_pct is not None:
        p.status = "successful" if p.collected_pct >= 50 else "failed"

    # Дата парсинга и флаг завершённости
    from datetime import date as _date
    p.parsed_date  = _date.today().isoformat()

    # dt/dd мета-поля
    for dt_el in soup.find_all("dt"):
        label = dt_el.get_text(strip=True).lower()
        dd_el = dt_el.find_next_sibling("dd")
        if not dd_el:
            continue
        val = dd_el.get_text(strip=True)
        if "поддержали" in label:
            p.backers_count = _int(val)
        elif "запущен" in label:
            raw_d = val.strip()
            if re.match(r"\d{2}\.\d{2}\.\d{4}", raw_d):
                p.date_start = raw_d
            else:
                months = {"января":"01","февраля":"02","марта":"03","апреля":"04",
                          "мая":"05","июня":"06","июля":"07","августа":"08",
                          "сентября":"09","октября":"10","ноября":"11","декабря":"12"}
                parts = raw_d.lower().split()
                if len(parts) == 2 and parts[1] in months:
                    from datetime import date as _dt2
                    p.date_start = f"{parts[0].zfill(2)}.{months[parts[1]]}.{_dt2.today().year}"
                else:
                    p.date_start = raw_d
        elif any(w in label for w in ("завершён","завершился","действовал до","окончание")):
            raw_end = val.strip()
            if re.match(r"\d{2}\.\d{2}\.\d{4}", raw_end):
                p.date_end = raw_end
        elif "осталось" in label:
            pass  # дней осталось — не сохраняем (меняется каждый день)
                  # для завершённых: duration = date_end - date_start (feature_engineering)
        elif "регион" in label:
            p.region = val
        elif "категория" in label:
            lnk = dd_el.find("a")
            if lnk:
                p.category = lnk.get_text(strip=True)

    # duration_days вычисляется в feature_engineering.py из date_end - date_start

    # Автор
    author_el = soup.find("h2", class_=re.compile("authorName"))
    p.author_name = author_el.get_text(strip=True) if author_el else None
    meta_vals = soup.find_all("dd", class_=re.compile("authorMetaValue"))
    if len(meta_vals) >= 1:
        p.author_projects = _int(meta_vals[0].get_text())
    if len(meta_vals) >= 2:
        p.author_backers  = _int(meta_vals[1].get_text())

    # Вознаграждения
    rewards = [r for r in soup.find_all("article", class_=re.compile("reward"))
               if r.get("id", "") != "reward-donate"]
    p.nmb_reward = len(rewards)
    prices = []
    for rw in rewards:
        price_el = rw.find("dd", class_=re.compile("priceValue"))
        if price_el:
            v = _money(price_el.get_text())
            if v is not None:
                prices.append(v)
    if prices:
        p.min_reward = min(prices)
        p.max_reward = max(prices)

    # Вкладки
    p.updates_count  = 0
    p.comments_count = 0
    for a in soup.select("a[class*='link']"):
        href_a = a.get("href", "")
        count_el = a.find("div", class_=re.compile("count"))
        count = _int(count_el.find("span").get_text()) if count_el and count_el.find("span") else 0
        if "/updates" in href_a:
            p.updates_count = count
        elif "/comments" in href_a:
            p.comments_count = count

    # Медиа и текст
    desc_block = soup.find("div", class_=re.compile(
        r"common-wrapper-module__wrapper|campaign-info-module__text"
    ))
    if desc_block:
        iframes = desc_block.find_all("iframe", src=re.compile(r"youtube|vimeo|rutube", re.I))
        video_tags = desc_block.find_all("video")
        p.nmb_video = len(iframes) + len(video_tags)
        imgs = [img for img in desc_block.find_all("img")
                if "fluentui-emoji" not in img.get("src", "")]
        p.nmb_image = len(imgs)
        raw_text = desc_block.get_text(" ", strip=True)
        p.nmb_word     = _words(raw_text)
        p.nmb_sentence = _sentences(raw_text)
        p.full_text = clean_text(raw_text) if raw_text else None
        # water_pct, spam_pct, text_tone — вычисляются в feature_engineering.py

    return p


# ─── Основной цикл ────────────────────────────────────────────────────────
def run():
    # 1. Собираем URL из всех Excel-файлов
    urls = extract_failed_urls(EXCEL_FILES)
    if not urls:
        log.error("Не найдено ни одного URL. Проверьте названия файлов в EXCEL_FILES.")
        return

    # 2. Загружаем уже собранные (для возобновления)
    output_path = Path(OUTPUT_FILE)
    seen_urls: set[str] = set()
    import csv as _csv2
    col_names = list(Project.__dataclass_fields__.keys())

    if output_path.exists():
        # Читаем файл и проверяем есть ли заголовок
        with open(output_path, encoding="utf-8-sig", newline="") as _f:
            reader = _csv2.reader(_f)
            all_rows = list(reader)

        if not all_rows:
            # Файл пустой — пишем заголовок
            pd.DataFrame(columns=col_names).to_csv(
                output_path, index=False, encoding="utf-8-sig", quoting=csv.QUOTE_ALL)
        else:
            first_row = all_rows[0]
            has_header = (first_row[0] == "project_id" if first_row else False)

            if not has_header:
                # Нет заголовка — восстанавливаем файл с заголовком
                log.warning("Заголовок отсутствует — восстанавливаем...")
                valid = [r for r in all_rows if len(r) == len(col_names)]
                df_fix = pd.DataFrame(valid, columns=col_names)
                df_fix.to_csv(output_path, index=False,
                              encoding="utf-8-sig", quoting=csv.QUOTE_ALL)
                log.info("Файл восстановлен: %d строк.", len(df_fix))
                # Перечитываем
                with open(output_path, encoding="utf-8-sig", newline="") as _f:
                    reader = _csv2.reader(_f)
                    all_rows = list(reader)

            # Собираем seen_urls (пропускаем строку заголовка)
            url_idx = all_rows[0].index("url") if "url" in all_rows[0] else 1
            for row in all_rows[1:]:
                if len(row) > url_idx and row[url_idx].startswith("http"):
                    seen_urls.add(_norm_url(row[url_idx]))
        log.info("Уже собрано: %d проектов.", len(seen_urls))
    else:
        # Новый файл — создаём с заголовком
        pd.DataFrame(columns=col_names).to_csv(
            output_path, index=False, encoding="utf-8-sig", quoting=csv.QUOTE_ALL)
        log.info("Создан файл: %s", output_path.absolute())

    to_parse = [u for u in urls if u not in seen_urls]
    log.info("Осталось распарсить: %d из %d", len(to_parse), len(urls))

    if not to_parse:
        log.info("Все URL уже собраны!")
        return

    # 3. Парсим
    driver  = make_driver()
    buffer  = []
    success = 0
    failed  = 0

    try:
        for i, url in enumerate(to_parse, 1):
            log.info("[%d/%d] %s", i, len(to_parse), url)
            project = parse_project(driver, url)

            if project:
                # Финальная проверка: убеждаемся что проект действительно неуспешный
                pct = project.collected_pct or 0
                if pct >= 50:
                    log.info("  SKIP (pct=%.0f%% >= 50, не неуспешный): '%s'",
                             pct, project.title)
                    seen_urls.add(url)
                    continue

                row = asdict(project)
                row["target"] = 0
                row["status"] = "failed"   # явно ставим статус
                buffer.append(row)
                seen_urls.add(url)
                success += 1
                log.info("  OK: '%s' | %.0f%% | target=0",
                         project.title, pct)
            else:
                failed += 1
                log.warning("  FAIL: %s", url)

            if len(buffer) >= 10:
                _flush(buffer, output_path)
                buffer.clear()
                log.info("  Сохранено. Итого: %d OK, %d FAIL", success, failed)

            time.sleep(random.uniform(DELAY_MIN, DELAY_MAX))

    except KeyboardInterrupt:
        log.info("Остановлено вручную.")
    finally:
        if buffer:
            _flush(buffer, output_path)
        driver.quit()

    # 4. Итог
    df = pd.read_csv(output_path, engine="python", on_bad_lines="warn")
    log.info("ИТОГО в %s: %d строк", output_path.name, len(df))
    print(f"\n{'='*50}")
    print(f"Неуспешных проектов собрано: {len(df)}")
    print(f"Успешно: {success} | Не удалось: {failed}")
    print(f"Файл: {output_path.absolute()}")
    if not df.empty and "title" in df.columns:
        print(df[["title", "collected_pct", "status"]].head(5).to_string())


def _flush(rows: list[dict], path: Path):
    pd.DataFrame(rows).to_csv(
        path, mode="a", header=False, index=False,
        encoding="utf-8-sig", quoting=csv.QUOTE_ALL
    )


if __name__ == "__main__":
    run()